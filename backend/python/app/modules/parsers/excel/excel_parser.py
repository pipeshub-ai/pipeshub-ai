import asyncio
import io
import json
import os
import re
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple, Union

from langchain_core.language_models.chat_models import BaseChatModel
from langchain_core.messages import AIMessage, HumanMessage
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from tenacity import (
    retry,
    stop_after_attempt,
    wait_exponential,
)

from app.models.blocks import (
    Block,
    BlockContainerIndex,
    BlockGroup,
    BlocksContainer,
    BlockType,
    DataFormat,
    GroupType,
    TableMetadata,
)
from app.modules.parsers.excel.prompt_template import (
    ExcelHeaderDetection,
    RowDescriptions,
    TableHeaders,
    excel_header_generation_prompt,
    excel_header_detection_prompt,
    prompt,
    row_text_prompt,
    sheet_summary_prompt,
    table_summary_prompt,
)
from app.utils.indexing_helpers import generate_simple_row_text
from app.utils.streaming import invoke_with_structured_output_and_reflection

# Module-level constants for Excel processing (mirror CSV parser)
NUM_SAMPLE_ROWS = 5  # Number of representative sample rows to select for header generation
MAX_HEADER_GENERATION_ROWS = 10  # Maximum number of rows to use for header generation
MAX_HEADER_DETECTION_ROWS = 4  # Check first 4 rows for multi-row headers
MIN_ROWS_FOR_HEADER_ANALYSIS = 2  # Minimum number of rows required for header analysis
MAX_HEADER_COUNT_RETRIES = 2  # Maximum retries when LLM returns wrong header count


# Built-in Excel date format codes mapping
BUILTIN_DATE_FORMATS = {
    14: "mm/dd/yyyy",
    15: "d-mmm-yy",
    16: "d-mmm",
    17: "mmm-yy",
    18: "h:mm AM/PM",
    19: "h:mm:ss AM/PM",
    20: "h:mm",
    21: "h:mm:ss",
    22: "m/d/yy h:mm",
}


def format_excel_datetime(dt_value: Any, number_format: str) -> Any:
    """
    Apply Excel number format to datetime value.
    
    Converts Python datetime objects to formatted strings matching Excel display format.
    Falls back to ISO format for unrecognized formats.
    
    Args:
        dt_value: The cell value (may or may not be a datetime)
        number_format: The Excel number format code (e.g., "mmm-yyyy", "dd-mmm-yy")
    
    Returns:
        Formatted string if datetime with valid format, otherwise original value
    """
    if not isinstance(dt_value, datetime):
        return dt_value
    
    if not number_format or number_format == "General":
        # No specific format, use ISO format as fallback
        return dt_value.isoformat()
    
    # Convert Excel format to Python strftime format
    try:
        # Handle built-in numeric format codes
        if number_format.isdigit():
            format_code = int(number_format)
            if format_code in BUILTIN_DATE_FORMATS:
                number_format = BUILTIN_DATE_FORMATS[format_code]
        
        # Determine if format contains time components
        has_hour = bool(re.search(r'(?<![a-z])h+(?![a-z])', number_format, flags=re.IGNORECASE))
        has_colon = ':' in number_format
        has_am_pm = 'am/pm' in number_format.lower() or 'a/p' in number_format.lower()
        has_time = has_hour or has_am_pm
        
        python_format = number_format
        
        # NEW STRATEGY: Pre-process to identify which 'mm' and 'm' tokens are in time vs date context
        # Mark time-context mm/m tokens BEFORE doing any replacements
        
        # Find position of first hour indicator (h or hh) to split date/time sections
        hour_pattern = re.search(r'(?<![a-z])h+(?![a-z])', python_format, flags=re.IGNORECASE)
        hour_start_pos = hour_pattern.start() if hour_pattern else len(python_format)
        
        # Mark all 'mm' and 'm' positions and their contexts (date vs time)
        def replace_mm_and_m_contextually(format_str, hour_boundary):
            """Replace mm and m based on their position relative to hour marker."""
            result = []
            i = 0
            while i < len(format_str):
                # Check for 'mmmm' (full month name) - handle first
                if format_str[i:i+4].lower() == 'mmmm':
                    result.append('mmmm')  # Keep for later replacement
                    i += 4
                # Check for 'mmm' (abbreviated month name)
                elif format_str[i:i+3].lower() == 'mmm':
                    result.append('mmm')  # Keep for later replacement
                    i += 3
                # Check for 'mm'
                elif format_str[i:i+2] == 'mm':
                    # Determine context: before/after hour marker, or near colons
                    # 1. Check if adjacent to colon (with optional spaces)
                    has_colon_before = (i > 0 and format_str[i-1:i].strip() == '' and 
                                       i > 1 and format_str[i-2] == ':') or (i > 0 and format_str[i-1] == ':')
                    has_colon_after = (i+2 < len(format_str) and format_str[i+2:i+3].strip() == '' and
                                      i+3 < len(format_str) and format_str[i+3] == ':') or (i+2 < len(format_str) and format_str[i+2] == ':')
                    
                    # 2. Check if in time section (after hour marker)
                    in_time_section = i >= hour_boundary
                    
                    # 3. Decide: minute or month
                    if has_colon_before or has_colon_after or (in_time_section and has_time):
                        result.append('__MINUTE_MM__')  # Placeholder for minute
                    elif has_colon and not has_hour:
                        # Special case: "mm:ss" format (no hours) - mm is minutes
                        result.append('__MINUTE_MM__')
                    else:
                        result.append('__MONTH_MM__')  # Placeholder for month
                    i += 2
                # Check for single 'm' (not part of mmm or mmmm)
                elif (format_str[i] == 'm' and 
                      (i == 0 or format_str[i-1:i+1].lower() not in ['mm', 'am']) and
                      (i+1 >= len(format_str) or format_str[i:i+2].lower() not in ['mm', 'mp'])):
                    # Similar logic for single 'm'
                    has_colon_before = i > 0 and format_str[i-1] == ':'
                    in_time_section = i >= hour_boundary
                    
                    if has_colon_before or (in_time_section and has_time):
                        result.append('__MINUTE_M__')  # Placeholder for minute
                    else:
                        result.append('__MONTH_M__')  # Placeholder for month
                    i += 1
                else:
                    result.append(format_str[i])
                    i += 1
            
            return ''.join(result)
        
        # Apply contextual marking
        python_format = replace_mm_and_m_contextually(python_format, hour_start_pos)
        # DEBUG
        import os
        if os.getenv('DEBUG_FORMAT'):
            print(f"DEBUG: After marking: '{python_format}'")
        
        # Now do standard replacements in order
        # 1. Handle full month/day names
        python_format = python_format.replace("mmmm", "%B")
        python_format = python_format.replace("dddd", "%A")
        
        # 2. Handle abbreviated names
        python_format = python_format.replace("mmm", "%b")
        python_format = python_format.replace("ddd", "%a")
        
        # 3. Handle years
        python_format = python_format.replace("yyyy", "%Y")
        python_format = python_format.replace("yy", "%y")
        
        # 4. Handle AM/PM
        python_format = re.sub(r'AM/PM', '%p', python_format, flags=re.IGNORECASE)
        python_format = re.sub(r'A/P', '%p', python_format, flags=re.IGNORECASE)
        
        # 5. Handle hours
        if 'hh' in python_format.lower():
            if has_am_pm:
                python_format = re.sub(r'hh', '%I', python_format, flags=re.IGNORECASE)
            else:
                python_format = re.sub(r'hh', '%H', python_format, flags=re.IGNORECASE)
        
        if re.search(r'(?<!%)h(?![a-zA-Z])', python_format, flags=re.IGNORECASE):
            if has_am_pm:
                python_format = re.sub(r'(?<!%)h(?![a-zA-Z])', '%I', python_format, flags=re.IGNORECASE)
            else:
                python_format = re.sub(r'(?<!%)h(?![a-zA-Z])', '%H', python_format, flags=re.IGNORECASE)
        
        # 6. Handle seconds
        python_format = python_format.replace("ss", "%S")
        python_format = re.sub(r'(?<!%)s(?![a-zA-Z])', '%S', python_format)
        
        # 7. Replace placeholders for mm and m
        python_format = python_format.replace('__MONTH_MM__', '%m')
        python_format = python_format.replace('__MINUTE_MM__', '%M')
        python_format = python_format.replace('__MONTH_M__', '%m')
        python_format = python_format.replace('__MINUTE_M__', '%M')
        
        # 8. Handle 'dd' - day with leading zero
        python_format = python_format.replace("dd", "%d")
        
        # 9. Handle single 'd' - day without leading zero
        if re.search(r'(?<!%)d(?![a-zA-Z])', python_format):
            python_format = re.sub(r'(?<!%)d(?![a-zA-Z])', '%d', python_format)
        
        # Apply the format
        formatted_value = dt_value.strftime(python_format)
        
        # Post-process to handle format nuances
        # 1. Strip leading zeros from single-digit representations
        
        # Check if original format had single 'd' (not 'dd')
        if re.search(r'(?<!d)d(?!d)', number_format.lower()):
            # Strip leading zero from day (e.g., "05" -> "5")
            # Match day numbers with leading zero (01-09) at word boundaries
            formatted_value = re.sub(r'(?<![0-9])0([1-9])(?=[^0-9]|$)', r'\1', formatted_value)
        
        # Check if original format had single 'm' for month (not 'mm', not in time context)
        if not has_time and re.search(r'(?<!m)m(?!m)', number_format.lower()):
            # Strip leading zero from month (e.g., "01" -> "1")
            formatted_value = re.sub(r'(?<![0-9])0([1-9])(?=[^0-9]|$)', r'\1', formatted_value)
        
        # Check if original format had single 'h' for hours (not 'hh')
        if re.search(r'(?<!h)h(?!h)', number_format.lower()):
            # Strip leading zero from hours (e.g., "02" -> "2")
            # This handles both 12-hour and 24-hour formats
            formatted_value = re.sub(r'(?<![0-9])0([1-9])(?=:)', r'\1', formatted_value)
        
        # Excel typically uses lowercase for abbreviated months in formats like 'mmm-yyyy'
        # Check if original format has lowercase 'mmm' (3-letter month)
        if 'mmm' in number_format.lower() and 'mmmm' not in number_format.lower():
            # Find and replace month abbreviations with lowercase
            # Python's %b produces capitalized month names (Jan, Feb, etc.)
            month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            for month_name in month_names:
                if month_name in formatted_value:
                    formatted_value = formatted_value.replace(month_name, month_name.lower(), 1)
                    break
        
        return formatted_value
        
    except Exception:
        # If formatting fails, fall back to ISO format
        return dt_value.isoformat()


class ExcelParser:
    def __init__(self, logger) -> None:
        self.logger = logger
        self.workbook = None
        self.file_binary = None

        # Store prompts
        self.sheet_summary_prompt = sheet_summary_prompt
        self.table_summary_prompt = table_summary_prompt
        self.row_text_prompt = row_text_prompt

        # Configure retry parameters
        self.max_retries = 3
        self.min_wait = 1  # seconds
        self.max_wait = 10  # seconds

    def load_workbook_from_binary(self, file_binary: bytes) -> None:
        """Load workbook from binary (no LLM calls).

        This is the first phase of Excel processing - pure parsing without LLM calls.
        """
        self.logger.info("Loading workbook from binary data")
        self.file_binary = file_binary
        if self.file_binary:
            self.workbook = load_workbook(io.BytesIO(file_binary), data_only=True)
            self.logger.info(f"Workbook loaded successfully with {len(self.workbook.sheetnames)} sheets: {self.workbook.sheetnames}")

    async def create_blocks(self, llm: BaseChatModel) -> BlocksContainer:
        """Create blocks from loaded workbook (involves LLM calls).

        This is the second phase - involves LLM calls for table summaries and row descriptions.
        Must call load_workbook_from_binary() first.
        """
        self.logger.info("Starting block creation from workbook with LLM")
        try:
            result = await self.get_blocks_from_workbook(llm)
            self.logger.info(f"Block creation completed. Generated {len(result.blocks)} blocks and {len(result.block_groups)} block groups")
            return result
        finally:
            if self.workbook:
                self.logger.info("Closing workbook")
                self.workbook.close()

    async def parse(self, file_binary: bytes, llm: BaseChatModel) -> BlocksContainer:
        """
        Parse Excel file and extract all content including sheets, cells, formulas, etc.

        For new code, prefer using load_workbook_from_binary() followed by create_blocks()
        to allow yielding progress events between phases.

        Returns:
            Dict containing parsed content with structure:
            {
                'sheets': List[Dict],        # List of sheet data
                'metadata': Dict,            # Workbook metadata
                'text_content': str,         # All text content concatenated
                'sheet_names': List[str],    # List of sheet names
                'total_rows': int,           # Total rows across all sheets
                'total_cells': int           # Total cells with content
            }
        """
        self.logger.info("Starting Excel file parsing")
        try:
            self.load_workbook_from_binary(file_binary)
            result = await self.get_blocks_from_workbook(llm)
            self.logger.info(f"Excel file parsing completed successfully")
            return result

        except Exception as e:
            self.logger.error(f"Error parsing Excel file: {e}", exc_info=True)
            raise
        finally:
            if self.workbook:
                self.logger.info("Closing workbook")
                self.workbook.close()

    def _json_default(self, obj) -> str:
        if isinstance(obj, datetime):
            return obj.isoformat()
        return str(obj)

    def _process_sheet(self, sheet) -> Dict[str, List[List[Dict[str, Any]]]]:
        """Process individual sheet and extract cell data"""
        try:
            self.logger.info(f"Processing sheet: {sheet.title}")
            sheet_data = {"headers": [], "data": []}

            # Extract headers from first row
            first_row = next(sheet.iter_rows(min_row=1, max_row=1))
            sheet_data["headers"] = [cell.value for cell in first_row]
            self.logger.info(f"Extracted {len(sheet_data['headers'])} headers from first row")

            # Start from second row
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
                row_data = []

                for col_idx, cell in enumerate(row, 1):
                    # Handle merged cells
                    if isinstance(cell, MergedCell):
                        cell_data = {
                            "value": None,  # Merged cells don't contain values
                            "header": (
                                sheet_data["headers"][col_idx - 1]
                                if col_idx - 1 < len(sheet_data["headers"])
                                else None
                            ),
                            "row": row_idx,
                            "column": col_idx,
                            # Use utility function instead
                            "column_letter": get_column_letter(col_idx),
                            "coordinate": f"{get_column_letter(col_idx)}{row_idx}",
                            "data_type": "merged",
                            "style": {"font": {}, "fill": {}, "alignment": {}},
                        }
                    else:
                        cell_data = {
                            "value": cell.value,
                            "header": (
                                sheet_data["headers"][col_idx - 1]
                                if col_idx - 1 < len(sheet_data["headers"])
                                else None
                            ),
                            "row": row_idx,
                            "column": col_idx,
                            "column_letter": cell.column_letter,
                            "coordinate": cell.coordinate,
                            "data_type": cell.data_type,
                            "style": {
                                "font": {
                                    "bold": cell.font.bold,
                                    "italic": cell.font.italic,
                                    "size": cell.font.size,
                                    "color": (
                                        cell.font.color.rgb if cell.font.color else None
                                    ),
                                },
                                "fill": {
                                    "background_color": (
                                        cell.fill.start_color.rgb
                                        if cell.fill.start_color
                                        else None
                                    )
                                },
                                "alignment": {
                                    "horizontal": cell.alignment.horizontal,
                                    "vertical": cell.alignment.vertical,
                                },
                            },
                        }

                        # Add formula if present
                        if cell.data_type == "f":
                            cell_data["formula"] = cell.value

                    row_data.append(cell_data)

                sheet_data["data"].append(row_data)

            self.logger.info(f"Processed {len(sheet_data['data'])} data rows from sheet: {sheet.title}")
            return sheet_data

        except Exception as e:
            self.logger.error(f"Error processing sheet {sheet.title}: {e}", exc_info=True)
            raise

    async def find_tables(self, sheet, llm: BaseChatModel) -> List[Dict[str, Any]]:
        """Find and process all tables in a sheet with LLM-based header detection/generation"""
        try:
            self.logger.info(f"Finding tables in sheet: {sheet.title}")
            tables = []
            visited_cells = set()  # Track already processed cells

            async def get_table(start_row: int, start_col: int) -> Dict[str, Any]:
                """Extract a table starting from (start_row, start_col) with intelligent header detection."""
                self.logger.info(f"Extracting table starting at row={start_row}, col={start_col}")
                # Step 1: Find table boundaries (max_row, max_col)
                max_col = start_col
                for col in range(start_col, sheet.max_column + 1):
                    has_data = False
                    for r in range(start_row, sheet.max_row + 1):
                        cell = sheet.cell(row=r, column=col)
                        if cell.value is not None:
                            has_data = True
                            max_col = col
                            break
                    if not has_data:
                        break

                max_row = start_row
                for row in range(start_row+1, sheet.max_row + 1):
                    has_data = False
                    for col in range(start_col, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        if cell.value is not None:
                            has_data = True
                            max_row = row
                            break
                    if not has_data and row != start_row+1:
                        break
                
                # Step 1.5: Expand left to include additional columns
                for col in range(start_col - 1, 0, -1):  # Go left from start_col to column 1
                    has_data = False
                    for row in range(start_row, max_row + 1):  # Check within rectangular region
                        cell = sheet.cell(row=row, column=col)
                        if cell.value is not None:
                            has_data = True
                            start_col = col  # Update start_col to include this column
                            break
                    if not has_data:
                        break  # Found empty column, stop expanding left
                
                column_count = max_col - start_col + 1
                self.logger.info(f"Table boundaries: rows [{start_row}-{max_row}], cols [{start_col}-{max_col}], column_count={column_count}")

                # Step 2: Extract first few rows for header detection
                first_rows = []
                for row_idx in range(start_row, min(start_row + MAX_HEADER_DETECTION_ROWS, max_row + 1)):
                    row_values = []
                    for col in range(start_col, max_col + 1):
                        cell = sheet.cell(row=row_idx, column=col)
                        cell_data = self._process_cell(cell, None, row_idx, col)
                        row_values.append(cell_data["value"])
                    first_rows.append(row_values)

                # Step 3: Detect headers with LLM
                detection = await self.detect_excel_headers_with_llm(first_rows, llm)
                self.logger.info(f"Header detection result: has_headers={detection.has_headers}, num_header_rows={detection.num_header_rows}, confidence={detection.confidence}")

                # Step 4: Determine headers and data start row
                headers = []
                data_start_row = start_row

                if detection.has_headers and detection.num_header_rows == 1:
                    # Single row header - use directly
                    headers = first_rows[0] if first_rows else []
                    data_start_row = start_row + 1
                    self.logger.info(f"Using single-row headers: {headers}")
                    # Mark header cells as visited
                    for col in range(start_col, max_col + 1):
                        visited_cells.add((start_row, col))
                elif detection.has_headers and detection.num_header_rows > 1:
                    # Multi-row headers: concatenate them into single-row headers
                    multirow_headers = first_rows[:detection.num_header_rows]
                    data_start_row = start_row + detection.num_header_rows
                    self.logger.info(f"Multi-row headers detected ({detection.num_header_rows} rows), will concatenate into single-row headers")
                    
                    # Mark header cells as visited
                    for row_idx in range(start_row, start_row + detection.num_header_rows):
                        for col in range(start_col, max_col + 1):
                            visited_cells.add((row_idx, col))
                    
                    # Concatenate multi-row headers directly (no LLM)
                    headers = self._concatenate_multirow_headers(multirow_headers, column_count)
                    self.logger.info(f"Concatenated headers: {headers}")
                else:
                    # No headers: all rows are data, generate headers from data
                    data_start_row = start_row
                    sample_start = start_row
                    self.logger.info("No headers detected, will generate headers from data")

                    # Extract all rows for sampling
                    all_rows = []
                    for row_idx in range(sample_start, max_row + 1):
                        row_values = []
                        for col in range(start_col, max_col + 1):
                            cell = sheet.cell(row=row_idx, column=col)
                            cell_data = self._process_cell(cell, None, row_idx, col)
                            row_values.append(cell_data["value"])
                        all_rows.append(row_values)

                    # Select representative sample rows
                    sample_rows = self._select_representative_sample_rows(all_rows, MAX_HEADER_GENERATION_ROWS)
                    self.logger.info(f"Selected {len(sample_rows)} representative sample rows for header generation")
                    
                    # Generate headers with LLM
                    headers = await self.generate_excel_headers_with_llm(sample_rows, column_count, llm)
                    self.logger.info(f"Generated headers: {headers}")

                # Normalize headers to match column count
                if headers:
                    # Replace None values in existing headers
                    for i in range(len(headers)):
                        if headers[i] is None or (isinstance(headers[i], str) and not headers[i].strip()):
                            headers[i] = f"Column_{i + 1}"
                            
                    # Pad if too short
                    if len(headers) < column_count:
                        self.logger.info(f"Padding headers from {len(headers)} to {column_count}")
                        for i in range(len(headers) + 1, column_count + 1):
                            headers.append(f"Column_{i}")
                    
                    # Truncate if too long (edge case)
                    elif len(headers) > column_count:
                        self.logger.warning(f"Truncating headers from {len(headers)} to {column_count}")
                        headers = headers[:column_count]
                    
                    self.logger.info(f"Normalized headers ({len(headers)} total): {headers}")

                # Handle empty headers case
                if not headers or all(h is None for h in headers):
                    return {
                        "headers": [],
                        "data": [],
                        "start_row": start_row,
                        "start_col": start_col,
                        "end_row": start_row,
                        "end_col": start_col,
                    }

                # Step 5: Build table structure once with correct headers
                table_data = []
                for row_idx in range(data_start_row, max_row + 1):
                    row_data = []
                    for col_idx, col in enumerate(range(start_col, max_col + 1)):
                        cell = sheet.cell(row=row_idx, column=col)
                        header = headers[col_idx]
                        cell_data = self._process_cell(cell, header, row_idx, col)
                        if cell.value is not None:
                            visited_cells.add((row_idx, col))
                        row_data.append(cell_data)
                    table_data.append(row_data)

                return {
                    "headers": headers,
                    "data": table_data,
                    "start_row": start_row,
                    "start_col": start_col,
                    "end_row": max_row,
                    "end_col": max_col,
                }

            # Find all tables in the sheet
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)

                    # Possible table start detection (cell has data and not visited)
                    if (
                        cell.value
                        and (row, col) not in visited_cells
                    ):
                        self.logger.info(f"Found potential table start at ({row}, {col}) with value: {cell.value}")
                        table = await get_table(row, col)
                        if table["data"]:  # Only add if table has data
                            tables.append(table)
                            self.logger.info(f"Table added with {len(table['data'])} rows and {len(table['headers'])} columns")

            self.logger.info(f"Found {len(tables)} tables in sheet: {sheet.title}")
            return tables

        except Exception as e:
            self.logger.error(f"Error finding tables in sheet {sheet.title}: {e}", exc_info=True)
            raise

    def _process_cell(self, cell, header, row, col) -> Dict[str, Any]:
        """Process a single cell and return its data with denormalized merged cell values."""
        try:
            # Check if the cell is a merged cell
            if isinstance(cell, MergedCell):
                # Look for the merged range that contains this cell.
                merged_value = None
                for merged_range in cell.parent.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Get the top-left cell of the merged range
                        top_left_cell = cell.parent.cell(
                            row=merged_range.min_row, column=merged_range.min_col
                        )
                        merged_value = top_left_cell.value
                        # Apply datetime formatting if applicable
                        if isinstance(merged_value, datetime) and hasattr(top_left_cell, 'number_format'):
                            merged_value = format_excel_datetime(merged_value, top_left_cell.number_format)
                        break

                return {
                    "value": merged_value,  # Use the top-left cell's value (formatted if datetime)
                    "header": header,
                    "row": row,
                    "column": col,
                    "column_letter": get_column_letter(col),
                    "coordinate": f"{get_column_letter(col)}{row}",
                    "data_type": "merged",
                    "style": {"font": {}, "fill": {}, "alignment": {}},
                }

            # If not a merged cell, process normally.
            # Apply datetime formatting if the cell contains a datetime value
            cell_value = cell.value
            if isinstance(cell_value, datetime) and hasattr(cell, 'number_format'):
                cell_value = format_excel_datetime(cell_value, cell.number_format)
            
            return {
                "value": cell_value,  # Now contains formatted string for datetime values
                "header": header,
                "row": row,
                "column": col,
                "column_letter": cell.column_letter,
                "coordinate": cell.coordinate,
                "data_type": cell.data_type,
                "style": {
                    "font": {
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "size": cell.font.size,
                        "color": cell.font.color.rgb if cell.font.color else None,
                    },
                    "fill": {
                        "background_color": (
                            cell.fill.start_color.rgb if cell.fill.start_color else None
                        )
                    },
                    "alignment": {
                        "horizontal": cell.alignment.horizontal,
                        "vertical": cell.alignment.vertical,
                    },
                },
            }
        except Exception as e:
            self.logger.error(f"Error processing cell at ({row}, {col}): {e}", exc_info=True)
            raise

    def _count_empty_values(self, row: Dict[str, Any]) -> int:
        """Count the number of empty/None values in a row"""
        return sum(1 for value in row.values() if value is None or value == "")

    def _select_representative_sample_rows(
        self, data_rows: List[List[Any]], num_sample_rows: int = NUM_SAMPLE_ROWS
    ) -> List[Tuple[int, List[Any], int]]:
        """
        Select representative sample rows from data by prioritizing rows with fewer empty values.

        This method selects up to num_sample_rows rows, prioritizing:
        1. Perfect rows with no empty values (stops early if enough are found)
        2. Rows with the fewest empty values as fallback

        Args:
            data_rows: List of rows (each row is a list of values)
            num_sample_rows: Number of sample rows to select (default: NUM_SAMPLE_ROWS)

        Returns:
            List of tuples (row_index, row_list, empty_count) sorted by original index
        """
        selected_rows = []
        fallback_rows = []

        for idx, row in enumerate(data_rows):
            # Count empty values in this row
            empty_count = sum(1 for value in row if value is None or (isinstance(value, str) and value.strip() == ""))

            if empty_count == 0:
                # Perfect row with no empty values
                selected_rows.append((idx, row, empty_count))
                if len(selected_rows) >= num_sample_rows:
                    break  # Early stop - found enough perfect rows
            else:
                # Keep track of best non-perfect rows as fallback
                fallback_rows.append((idx, row, empty_count))

        # If we didn't find enough perfect rows, supplement with the best fallback rows
        if len(selected_rows) < num_sample_rows:
            # Sort fallback rows by empty count (ascending), then by index
            fallback_rows.sort(key=lambda x: (x[2], x[0]))
            # Add the best fallback rows to reach the target count
            needed = num_sample_rows - len(selected_rows)
            selected_rows.extend(fallback_rows[:needed])

        # Sort by original index to maintain logical order
        selected_rows.sort(key=lambda x: x[0])

        return selected_rows

    def _convert_rows_to_strings(self, rows: List[List[Any]], num_rows: int = 4) -> List[List[str]]:
        """
        Convert multiple rows to lists of strings for LLM prompts.

        Args:
            rows: List of rows where each row is a list of values
            num_rows: Number of rows to convert (default: 4)

        Returns:
            List of converted rows, where each row is a list of strings.
            Returns empty list for rows that don't exist.
        """
        return [
            [str(v) if v is not None else "" for v in rows[i]] if i < len(rows) else []
            for i in range(num_rows)
        ]

    async def detect_excel_headers_with_llm(
        self, first_rows: List[List[Any]], llm: BaseChatModel
    ) -> ExcelHeaderDetection:
        """
        Use LLM to detect if the first row(s) contain valid headers and how many rows they span.

        Args:
            first_rows: List of first few rows as lists of values (typically 4-6 rows)
            llm: Language model instance

        Returns:
            ExcelHeaderDetection with has_headers, num_header_rows, confidence, and reasoning
        """
        self.logger.info(f"Detecting headers with LLM for {len(first_rows)} rows")
        try:
            if len(first_rows) < MIN_ROWS_FOR_HEADER_ANALYSIS:
                self.logger.warning(f"Only {len(first_rows)} rows available, insufficient for header analysis (min: {MIN_ROWS_FOR_HEADER_ANALYSIS})")
                # Not enough rows to analyze, assume single-row headers exist
                return ExcelHeaderDetection(
                    has_headers=False,
                    num_header_rows=0,
                    confidence="low",
                    reasoning="Insufficient rows for analysis, defaulting to no headers"
                )

            # Prepare rows for prompt (convert to strings for display)
            row1, row2, row3, row4 = self._convert_rows_to_strings(first_rows, 4)
            self.logger.info("Calling LLM for header detection")
            messages = excel_header_detection_prompt.format_messages(
                row1=row1,
                row2=row2,
                row3=row3,
                row4=row4,
            )

            # Use centralized utility with reflection
            parsed_response = await invoke_with_structured_output_and_reflection(
                llm, messages, ExcelHeaderDetection
            )

            if parsed_response is not None:
                self.logger.info(f"LLM header detection successful: {parsed_response.reasoning}")
                # Validate num_header_rows is sensible
                if parsed_response.num_header_rows < 0:
                    parsed_response.num_header_rows = 0
                
                # If has_headers is True but num_header_rows is 0, correct it
                if parsed_response.has_headers and parsed_response.num_header_rows == 0:
                    parsed_response.num_header_rows = 1
                
                # If has_headers is False, ensure num_header_rows is 0
                if not parsed_response.has_headers:
                    parsed_response.num_header_rows = 0
                
                return parsed_response

            # Fallback: assume single-row headers exist if LLM fails
            self.logger.warning("Header detection LLM call failed, defaulting to no headers")
            return ExcelHeaderDetection(
                has_headers=False,
                num_header_rows=0,
                confidence="low",
                reasoning="LLM call failed, defaulting to no headers"
            )

        except Exception as e:
            self.logger.warning(f"Error in Excel header detection: {e}, defaulting to no headers")
            return ExcelHeaderDetection(
                has_headers=False,
                num_header_rows=0,
                confidence="low",
                reasoning=f"Error occurred: {str(e)}, defaulting to no headers"
            )

    async def generate_excel_headers_with_llm(
        self, sample_rows: List[Tuple[int, List[Any], int]], column_count: int, llm: BaseChatModel
    ) -> List[str]:
        """
        Generate descriptive headers from sample data using LLM.

        Used for two scenarios:
        1. No headers detected - generate from data rows
        2. Multi-row headers detected - generate from all rows including the multi-row headers

        Args:
            sample_rows: List of tuples (row_index, row_list, empty_count) from _select_representative_sample_rows
            column_count: Number of columns expected
            llm: Language model instance

        Returns:
            List of generated header names (always exactly column_count items)
        """
        self.logger.info(f"Generating headers with LLM for {column_count} columns using {len(sample_rows)} sample rows")
        
        try:
            # Format sample data for display
            formatted_samples = []
            for idx, row, empty_count in sample_rows[:MAX_HEADER_GENERATION_ROWS]:
                formatted_row = [str(v) if v is not None else "" for v in row]
                formatted_samples.append(formatted_row)

            # Format as JSON string for prompt
            sample_data_str = json.dumps(formatted_samples, indent=2)

            # Initial messages
            messages = excel_header_generation_prompt.format_messages(
                sample_data=sample_data_str,
                column_count=column_count,
                sample_count=len(formatted_samples),
            )

            # Retry loop for count mismatches
            for attempt in range(MAX_HEADER_COUNT_RETRIES + 1):
                if attempt > 0:
                    self.logger.info(f"Retry attempt {attempt}/{MAX_HEADER_COUNT_RETRIES} for header generation")
                else:
                    self.logger.info("Calling LLM for header generation (initial attempt)")

                # Use centralized utility with reflection for parse errors
                parsed_response = await invoke_with_structured_output_and_reflection(
                    llm, messages, TableHeaders
                )

                if parsed_response is not None and parsed_response.headers:
                    generated_headers = parsed_response.headers
                    self.logger.info(f"LLM generated {len(generated_headers)} headers (expected {column_count})")

                    # Validate header count matches column count
                    if len(generated_headers) == column_count:
                        self.logger.info(f"Successfully generated headers matching expected column count ({column_count})")
                        return generated_headers
                    else:
                        self.logger.warning(
                            f"Header count mismatch: generated {len(generated_headers)}, expected {column_count}. "
                            f"Headers: {generated_headers}"
                        )
                        
                        # If we have retries left, add reflection message for count correction
                        if attempt < MAX_HEADER_COUNT_RETRIES:
                            self.logger.info("Adding reflection message to correct header count")
                            
                            # Convert messages to list if not already
                            messages_list = list(messages)
                            
                            # Add the failed response to context
                            failed_response = json.dumps({"headers": generated_headers}, indent=2)
                            messages_list.append(AIMessage(content=failed_response))
                            
                            # Add reflection prompt
                            reflection_prompt = f"""Your previous response contained {len(generated_headers)} headers, but I need EXACTLY {column_count} headers.

Previous headers you provided: {generated_headers}

ERROR: You returned {len(generated_headers)} headers but the data has {column_count} columns.

Please correct your response:
- Analyze the sample data again carefully
- Count that there are {column_count} columns in the data
- Generate EXACTLY {column_count} headers, one for each column
- Verify your count before responding

Respond with ONLY a JSON object with EXACTLY {column_count} headers:
{{
    "headers": ["Header1", "Header2", ..., "Header{column_count}"]
}}"""
                            
                            messages_list.append(HumanMessage(content=reflection_prompt))
                            messages = messages_list
                            continue  # Try again
                        else:
                            # Out of retries, try smart fallback
                            self.logger.warning(f"Exhausted {MAX_HEADER_COUNT_RETRIES} retries, attempting smart adjustment")
                            adjusted_headers = self._adjust_headers_to_count(
                                generated_headers, column_count, formatted_samples
                            )
                            if adjusted_headers:
                                self.logger.info(f"Successfully adjusted headers to match count ({column_count})")
                                return adjusted_headers
                else:
                    self.logger.warning("LLM returned no response or empty headers")
                    break  # Exit retry loop

            # Final fallback: generate generic headers
            self.logger.warning(f"Using generic fallback headers for {column_count} columns")
            return [f"Column_{i}" for i in range(1, column_count + 1)]

        except Exception as e:
            self.logger.error(f"Error in Excel header generation: {e}", exc_info=True)
            self.logger.warning(f"Using generic fallback headers for {column_count} columns")
            return [f"Column_{i}" for i in range(1, column_count + 1)]


    def _adjust_headers_to_count(
        self, generated_headers: List[str], expected_count: int, sample_data: List[List[str]]
    ) -> Optional[List[str]]:
        """
        Intelligently adjust header count to match expected count.
        
        This is a smart fallback that tries to salvage LLM-generated headers when the count is close
        but not exact. Only works if the count difference is reasonable (within 20%).
        
        Args:
            generated_headers: Headers generated by LLM (wrong count)
            expected_count: Expected number of headers
            sample_data: Sample data rows to infer missing headers from
            
        Returns:
            Adjusted headers list matching expected_count, or None if adjustment not feasible
        """
        current_count = len(generated_headers)
        count_diff = abs(current_count - expected_count)
        
        # Only adjust if difference is reasonable (within 20% of expected)
        if count_diff > max(1, expected_count * 0.2):
            self.logger.info(
                f"Header count difference too large ({count_diff} headers, {count_diff/expected_count*100:.1f}% off), "
                "cannot adjust intelligently"
            )
            return None
            
        self.logger.info(
            f"Attempting to adjust {current_count} headers to {expected_count} "
            f"(difference: {count_diff}, {count_diff/expected_count*100:.1f}%)"
        )
        
        if current_count < expected_count:
            # Need to add headers (padding)
            adjusted = generated_headers.copy()
            num_to_add = expected_count - current_count
            
            self.logger.info(f"Padding {num_to_add} missing headers")
            
            # Try to infer names from sample data for missing columns
            for col_idx in range(current_count, expected_count):
                # Analyze the column data in samples to infer a good name
                if sample_data and col_idx < len(sample_data[0]):
                    column_values = [row[col_idx] if col_idx < len(row) else "" for row in sample_data]
                    # Filter out empty values
                    non_empty = [v for v in column_values if v and str(v).strip()]
                    
                    if non_empty:
                        # Try to infer type from data
                        sample_val = str(non_empty[0])
                        
                        # Check for common patterns
                        if re.match(r'^\d{4}-\d{2}-\d{2}', sample_val):
                            header = f"Date_{col_idx + 1}"
                        elif re.match(r'^[\$€£¥]?[\d,]+\.?\d*$', sample_val):
                            header = f"Amount_{col_idx + 1}"
                        elif re.match(r'^\d+\.?\d*%?$', sample_val):
                            header = f"Value_{col_idx + 1}"
                        elif re.match(r'^[A-Z]{2,}$', sample_val):
                            header = f"Code_{col_idx + 1}"
                        else:
                            header = f"Field_{col_idx + 1}"
                    else:
                        header = f"Column_{col_idx + 1}"
                else:
                    header = f"Column_{col_idx + 1}"
                
                adjusted.append(header)
                self.logger.info(f"Added inferred header for column {col_idx + 1}: {header}")
            
            return adjusted
            
        else:
            # Need to remove headers (truncating)
            self.logger.info(f"Truncating {current_count - expected_count} excess headers")
            adjusted = generated_headers[:expected_count]
            
            truncated = generated_headers[expected_count:]
            self.logger.info(f"Removed excess headers: {truncated}")
            
            return adjusted

    def _concatenate_multirow_headers(self, multirow_headers: List[List[Any]], column_count: int) -> List[str]:
        """
        Fallback method to concatenate multi-row headers with underscores.

        Args:
            multirow_headers: List of header rows
            column_count: Expected number of columns

        Returns:
            List of concatenated header strings
        """
        self.logger.info(f"Using simple concatenation for {len(multirow_headers)} header rows")
        consolidated = []
        
        for col_idx in range(column_count):
            # Collect non-empty values from all header rows for this column
            parts = []
            seen = set()  # Track seen values to avoid duplicates (e.g., from merged cells)
            
            for header_row in multirow_headers:
                if col_idx < len(header_row):
                    value = header_row[col_idx]
                    if value is not None and str(value).strip():
                        value_str = str(value).strip()
                        # Only add if we haven't seen this value already (handles merged cells)
                        if value_str not in seen:
                            parts.append(value_str)
                            seen.add(value_str)
            
            # Join with underscores or use generic name if no parts
            if parts:
                header = "_".join(parts)
            else:
                header = f"Column_{col_idx + 1}"
            
            consolidated.append(header)
        
        return consolidated

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=1, max=10),
        before_sleep=lambda retry_state: retry_state.args[0].logger.warning(
            f"Retrying LLM call after error. Attempt {retry_state.attempt_number}"
        ),
    )
    async def _call_llm(self,messages) -> Union[str, dict, list]:
        """Wrapper for LLM calls with retry logic"""
        return await self.llm.ainvoke(messages)

    async def get_tables_in_sheet(self, sheet_name: str, llm: BaseChatModel) -> List[Dict[str, Any]]:
        """Get all tables in a specific sheet with LLM-based header detection/generation
        
        Note: Header detection and generation is now handled in find_tables() method,
        so this method simply returns the tables with properly detected/generated headers.
        """
        self.logger.info(f"Getting tables in sheet: {sheet_name}")
        try:
            if not self.workbook:
                self.parse()

            if sheet_name not in self.workbook.sheetnames:
                self.logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                return []

            sheet = self.workbook[sheet_name]
            # find_tables now handles header detection/generation internally
            tables = await self.find_tables(sheet, llm)

            self.logger.info(f"Retrieved {len(tables)} tables from sheet: {sheet_name}")
            return tables

        except Exception as e:
            self.logger.error(f"Error getting tables in sheet {sheet_name}: {e}", exc_info=True)
            raise

    async def get_table_summary(self, table: Dict[str, Any]) -> str:
        """Get a natural language summary of a specific table"""
        self.logger.info(f"Getting summary for table with {len(table['headers'])} columns and {len(table['data'])} rows")
        try:
            # Prepare sample data
            sample_data = [
                {
                    cell["header"]: (
                        cell["value"].isoformat()
                        if isinstance(cell["value"], datetime)
                        else cell["value"]
                    )
                    for cell in row
                }
                for row in table["data"][:3]  # Use first 3 rows as sample
            ]

            # Get summary from LLM with retry
            messages = self.table_summary_prompt.format_messages(
                headers=table["headers"], sample_data=json.dumps(sample_data, indent=2)
            )
            response = await self._call_llm(messages)
            if '</think>' in response.content:
                response.content = response.content.split('</think>')[-1]
            self.logger.info(f"Table summary generated")
            return response.content

        except Exception as e:
            self.logger.error(f"Error getting table summary: {e}", exc_info=True)
            raise

    async def get_rows_text(
        self, rows: List[List[Dict[str, Any]]], table_summary: str
    ) -> List[str]:
        """Convert multiple rows into natural language text using context from summaries in a single prompt"""
        self.logger.info(f"Converting {len(rows)} rows to natural language text")
        try:
            # Prepare rows data
            rows_data = [
                {
                    cell["header"]: (
                        cell["value"].isoformat()
                        if isinstance(cell["value"], datetime)
                        else cell["value"]
                    )
                    for cell in row
                }
                for row in rows
            ]

            # Get natural language text from LLM with retry
            messages = self.row_text_prompt.format_messages(
                table_summary=table_summary, rows_data=json.dumps(rows_data, indent=2)
            )

            # Default to string representations of rows
            descriptions = [str(row) for row in rows_data]

            # Use centralized utility with reflection
            parsed_response = await invoke_with_structured_output_and_reflection(
                self.llm, messages, RowDescriptions
            )

            if parsed_response is not None and parsed_response.descriptions:
                descriptions = parsed_response.descriptions
                self.logger.info(f"Successfully generated natural language descriptions for {len(descriptions)} rows")

            return descriptions
        except Exception as e:
            self.logger.error(f"Error converting rows to natural language text: {e}", exc_info=True)
            raise

    async def process_sheet_with_summaries(
        self, llm, sheet_name: str, cumulative_row_count: List[int]
    ) -> Dict[str, Any]:
        """Process a sheet and generate all summaries and row texts
        Args:
            llm: Language model instance
            sheet_name: Name of the sheet to process
            cumulative_row_count: List with single element [count] to track cumulative rows across all tables
        """
        self.logger.info(f"Processing sheet with summaries: {sheet_name}")
        self.llm = llm

        if sheet_name not in self.workbook.sheetnames:
            self.logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            return None

        # Get threshold from environment variable (default: 1000)
        threshold = int(os.getenv("MAX_TABLE_ROWS_FOR_LLM", "1000"))
        self.logger.info(f"Using LLM threshold for row processing: {threshold} (cumulative count: {cumulative_row_count[0]})")

        # Get tables in the sheet
        tables = await self.get_tables_in_sheet(sheet_name, llm)

        # Process each table
        processed_tables = []
        for table_idx, table in enumerate(tables, 1):
            self.logger.info(f"Processing table {table_idx}/{len(tables)} in sheet {sheet_name}")
            # Get table summary (always use LLM)
            table_summary = await self.get_table_summary(table)

            # Add current table rows to cumulative count
            table_row_count = len(table["data"])
            cumulative_row_count[0] += table_row_count
            self.logger.info(f"Table has {table_row_count} rows, cumulative count: {cumulative_row_count[0]}")

            # Check if cumulative count exceeds threshold
            use_llm_for_rows = cumulative_row_count[0] <= threshold

            processed_rows = []

            if use_llm_for_rows:
                self.logger.info(f"Using LLM for row processing (under threshold of {threshold})")
                # Process rows in batches of 50 in parallel using LLM
                batch_size = 50

                # Create batches
                batches = []
                for i in range(0, len(table["data"]), batch_size):
                    batch = table["data"][i : i + batch_size]
                    batches.append((i, batch))  # Store start index and batch data

                self.logger.info(f"Processing {len(table['data'])} rows in {len(batches)} batches of {batch_size}")
                
                # Limit parallel processing to at most 10 concurrent batches
                semaphore = asyncio.Semaphore(10)

                async def limited_get_rows_text(batch) -> List[str]:
                    async with semaphore:
                        return await self.get_rows_text(batch, table_summary)

                # Create throttled tasks for all batches
                batch_tasks = []
                for start_idx, batch in batches:
                    task = limited_get_rows_text(batch)
                    batch_tasks.append((start_idx, batch, task))

                # Wait for all batches to complete (max 10 running concurrently)
                task_results = await asyncio.gather(*[task for _, _, task in batch_tasks])
                self.logger.info(f"Completed processing {len(batch_tasks)} batches with LLM")

                # Combine results with their metadata and process
                for i, (start_idx, batch, _) in enumerate(batch_tasks):
                    row_texts = task_results[i]

                    # Add processed rows to results
                    for row, row_text in zip(batch, row_texts):
                        if row:
                            processed_rows.append(
                                {
                                    "raw_data": {cell["header"]: cell["value"] for cell in row},
                                    "natural_language_text": row_text,
                                    "row_num": row[0]["row"],  # Include row number
                                }
                            )
            else:
                self.logger.info(f"Using simple format for row processing (exceeded threshold of {threshold})")
                # Use simple format for rows (skip LLM)
                for row in table["data"]:
                    if row:
                        row_data = {cell["header"]: cell["value"] for cell in row}
                        row_text = generate_simple_row_text(row_data)
                        processed_rows.append(
                            {
                                "raw_data": row_data,
                                "natural_language_text": row_text,
                                "row_num": row[0]["row"]  # Include row number
                            }
                        )

            processed_tables.append(
                {
                    "headers": table["headers"],
                    "summary": table_summary,
                    "rows": processed_rows,
                    "location": {
                        "start_row": table["start_row"],
                        "start_col": table["start_col"],
                        "end_row": table["end_row"],
                        "end_col": table["end_col"],
                    },
                }
            )
            self.logger.info(f"Completed processing table {table_idx} with {len(processed_rows)} rows")

        self.logger.info(f"Completed processing sheet {sheet_name} with {len(processed_tables)} tables")
        return {"sheet_name": sheet_name, "tables": processed_tables}

    async def get_blocks_from_workbook(self, llm) -> BlocksContainer:
        """Build a BlocksContainer with SHEET and TABLE groups and TABLE_ROW blocks.

        Mirrors the CSV blocks structure, but nests tables under sheet groups.
        """
        self.logger.info("Building blocks from workbook")
        blocks: List[Block] = []
        block_groups: List[BlockGroup] = []

        # Initialize cumulative row count for record-level threshold checking
        cumulative_row_count = [0]

        # Iterate sheets and build hierarchy
        self.logger.info(f"Processing {len(self.workbook.sheetnames)} sheets: {self.workbook.sheetnames}")
        for sheet_idx, sheet_name in enumerate(self.workbook.sheetnames, 1):
            self.logger.info(f"Processing sheet {sheet_idx}/{len(self.workbook.sheetnames)}: {sheet_name}")
            sheet_result = await self.process_sheet_with_summaries(llm, sheet_name, cumulative_row_count)
            if sheet_result is None:
                continue

            # Create SHEET group
            sheet_group_index = len(block_groups)
            sheet_group_children: List[BlockContainerIndex] = []
            sheet_group = BlockGroup(
                index=sheet_group_index,
                name=sheet_result["sheet_name"],
                type=GroupType.SHEET,
                parent_index=None,
                description=None,
                table_metadata=None,
                data={
                    "sheet_name": sheet_result["sheet_name"],
                    "table_count": len(sheet_result["tables"]),
                },
                format=DataFormat.JSON,
            )
            block_groups.append(sheet_group)

            # Add TABLE groups under this sheet
            for table in sheet_result["tables"]:
                table_group_index = len(block_groups)

                headers = table.get("headers", [])
                rows = table.get("rows", [])

                table_group_children: List[BlockContainerIndex] = []
                table_markdown = self.to_markdown(headers, rows)
                table_group = BlockGroup(
                    index=table_group_index,
                    name=None,
                    type=GroupType.TABLE,
                    parent_index=sheet_group_index,
                    description=None,
                    source_group_id=None,
                    table_metadata=TableMetadata(
                        num_of_rows=len(rows),
                        num_of_cols=len(headers) if headers else (len(rows[0]["raw_data"]) if rows else 0),
                    ),
                    data={
                        "table_summary": table.get("summary", ""),
                        "column_headers": headers,
                        "sheet_number": sheet_idx,
                        "sheet_name": sheet_name,
                        "table_markdown": table_markdown,
                    },
                    format=DataFormat.JSON,
                )
                block_groups.append(table_group)
                sheet_group_children.append(BlockContainerIndex(block_group_index=table_group_index))

                # Create TABLE_ROW blocks under this table
                for i, row in enumerate(rows):
                    block_index = len(blocks)
                    row_data = row.get("raw_data", {})
                    blocks.append(
                        Block(
                            index=block_index,
                            type=BlockType.TABLE_ROW,
                            format=DataFormat.JSON,
                            data={
                                "row_natural_language_text": row.get("natural_language_text", ""),
                                "row_number": int(row.get("row_num") or (i + 1)),
                                "row": json.dumps(row_data, default=self._json_default),
                                "sheet_number": sheet_idx,
                                "sheet_name": sheet_name,
                            },
                            parent_index=table_group_index,
                        )
                    )
                    table_group_children.append(BlockContainerIndex(block_index=block_index))

                # attach table children
                block_groups[table_group_index].children = table_group_children

            # attach sheet children (its tables)
            block_groups[sheet_group_index].children = sheet_group_children
            self.logger.info(f"Completed processing sheet {sheet_name}: {len(sheet_result['tables'])} tables")

        self.logger.info(f"Workbook processing complete. Total: {len(blocks)} blocks, {len(block_groups)} block groups")
        return BlocksContainer(blocks=blocks, block_groups=block_groups)

    def to_markdown(self, headers: List[str], rows: List[Dict[str, Any]]) -> str:
            """
            Convert CSV data to markdown table format.
            Args:
                data: List of dictionaries from read_stream() method
            Returns:
                String containing markdown formatted table
            """
            if not headers and not rows:
                return ""

            # Get headers from the first row
            headers = list(headers)

            # Start building the markdown table
            markdown_lines = []

            # Add header row
            header_row = "| " + " | ".join(str(header) for header in headers) + " |"
            markdown_lines.append(header_row)

            # Add separator row
            separator_row = "|" + "|".join(" --- " for _ in headers) + "|"
            markdown_lines.append(separator_row)
            data = []
            for row in rows:
                data.append(row.get("raw_data", {}))
            # Add data rows
            for row in data:
                # Handle None values and convert to string, escape pipe characters
                formatted_values = []
                for header in headers:
                    value = row.get(header, "")
                    if value is None:
                        value = ""
                    # Escape pipe characters and convert to string
                    value_str = str(value).replace("|", "\\|")
                    formatted_values.append(value_str)

                data_row = "| " + " | ".join(formatted_values) + " |"
                markdown_lines.append(data_row)

            return "\n".join(markdown_lines)

