"""
Parse file bytes to text content.

Supports PDF, OOXML (docx, xlsx, pptx), OLE2 (doc, xls, ppt),
and plain text (json, html, md, csv, code files, etc.).
Uses UTF-8 and Latin-1 only for text decoding.

A file extension is always required. Unsupported or missing extensions
are rejected immediately before any bytes are read.
"""

import json
from io import BytesIO
from typing import Optional

import fitz
from openpyxl import load_workbook

from app.modules.parsers.docx.docparser import DocParser
from app.modules.parsers.docx.docx_parser import DocxParser
from app.modules.parsers.excel.xls_parser import XLSParser
from app.modules.parsers.pptx.ppt_parser import PPTParser
from app.modules.parsers.pptx.pptx_parser import PPTXParser


# ---------------------------------------------------------------------------
# Supported extensions grouped by the parser they route to.
# These sets are the single source of truth — add/remove extensions here only.
# ---------------------------------------------------------------------------

SUPPORTED_PDF: frozenset[str] = frozenset({"pdf"})

SUPPORTED_OOXML: frozenset[str] = frozenset({"docx", "xlsx", "pptx"})

SUPPORTED_OLE2: frozenset[str] = frozenset({"doc", "xls", "ppt"})

SUPPORTED_TEXT: frozenset[str] = frozenset({
    # Documents / markup
    "txt", "md", "markdown", "rst", "tex",
    "html", "htm", "xhtml",
    "xml", "svg",
    # Data interchange
    "json", "jsonl", "ndjson",
    "csv", "tsv",
    "yaml", "yml", "toml",
    # Config / env
    "ini", "cfg", "conf", "env", "properties",
    # Notebooks
    "ipynb",
    # Code
    "py", "pyw",
    "js", "mjs", "cjs", "ts", "jsx", "tsx",
    "java", "kt", "kts",
    "c", "h", "cpp", "cc", "cxx", "hpp",
    "cs", "fs", "vb",
    "go", "rs", "rb", "php", "swift",
    "r", "rmd",
    "sh", "bash", "zsh", "fish", "ps1", "bat", "cmd",
    "sql", "graphql", "proto",
    "lua", "pl", "pm", "scala", "groovy", "dart", "ex", "exs",
    # Logs / misc
    "log", "diff", "patch",
})

# Master allow-list — union of all groups
ALL_SUPPORTED: frozenset[str] = (
    SUPPORTED_PDF | SUPPORTED_OOXML | SUPPORTED_OLE2 | SUPPORTED_TEXT
)


class FileContentParser:
    """Parse file bytes to text content.

    The file extension is **required** and acts as the primary routing key.
    Unsupported or missing extensions are rejected before any bytes are read.
    Magic bytes are used as a safety check for binary formats to catch
    mismatched or corrupted files early.

    Usage::

        parser = FileContentParser()
        ok, result = parser.parse(raw_bytes, ext="docx")
    """

    # Magic-byte constants — used only for safety validation, not routing.
    _MAGIC_PDF  = b"%PDF"
    _MAGIC_ZIP  = b"PK\x03\x04"
    _MAGIC_OLE2 = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

    def __init__(self) -> None:
        pass

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------

    def parse(
        self,
        raw: bytes,
        ext: str,
        max_bytes: Optional[int] = 500_000,
    ) -> tuple[bool, str]:
        """Parse file bytes to text content.

        Args:
            raw:       File content as bytes.
            ext:       File extension, required (with or without leading dot,
                       e.g. ``"pdf"`` or ``".pdf"``).  If the extension is
                       missing or not in the supported list the call returns
                       immediately with an error — no bytes are inspected.
            max_bytes: Optional cap on extracted text size (bytes). Content is
                       truncated beyond this limit.

        Returns:
            ``(success, json_string)``

            On success::

                {
                  "content":    str,   # extracted text
                  "truncated":  bool,  # True when content was cut
                  "bytes_read": int,   # original file size in bytes
                  "format":     str    # format label (mirrors the extension)
                }

            On failure::

                {"error": "<reason>"}
        """
        ext = ext.lower().lstrip(".") if ext else ""
        original_size = len(raw)

        # --- 1. Extension gate: reject anything missing or unsupported --------
        if not ext:
            return False, json.dumps({"error": "File extension is required"})
        if ext not in ALL_SUPPORTED:
            return False, json.dumps({"error": f"Unsupported file type: .{ext}"})

        # --- 2. Safety check: magic bytes must match what the extension claims -
        magic_error = self._validate_magic(raw, ext)
        if magic_error:
            return False, json.dumps({"error": magic_error})

        # --- 3. Route directly by extension — no further detection needed ------
        if ext in SUPPORTED_PDF:
            return self._parse_pdf(raw, original_size, max_bytes)
        if ext in SUPPORTED_OOXML:
            return self._parse_ooxml(raw, original_size, max_bytes, ext)
        if ext in SUPPORTED_OLE2:
            return self._parse_ole2(raw, original_size, max_bytes, ext)
        # SUPPORTED_TEXT — everything else
        return self._parse_plain_text(raw, original_size, max_bytes, ext)

    # ------------------------------------------------------------------
    # Magic-byte safety validation
    # ------------------------------------------------------------------

    def _validate_magic(self, data: bytes, ext: str) -> Optional[str]:
        """Return an error string if magic bytes contradict the extension,
        or ``None`` if everything looks consistent.

        Text-based formats are not validated (they have no magic bytes).
        """
        if ext in SUPPORTED_PDF:
            if data[:4] != self._MAGIC_PDF:
                return f"File does not look like a PDF (unexpected magic bytes)"

        elif ext in SUPPORTED_OOXML:
            if data[:4] != self._MAGIC_ZIP:
                return f".{ext} file does not look like a ZIP/OOXML container"

        elif ext in SUPPORTED_OLE2:
            if data[:8] != self._MAGIC_OLE2:
                return f".{ext} file does not look like an OLE2 container"

        return None

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _truncate_text(
        self, text: str, max_bytes: Optional[int], encoding: str = "utf-8"
    ) -> tuple[str, bool]:
        """Truncate *text* so its encoded form fits within *max_bytes*."""
        if not max_bytes:
            return text, False
        encoded = text.encode(encoding)
        if len(encoded) <= max_bytes:
            return text, False
        return encoded[:max_bytes].decode(encoding, errors="ignore"), True

    def _decode_bytes(self, raw: bytes) -> str:
        """Decode bytes as UTF-8, falling back to Latin-1."""
        try:
            return raw.decode("utf-8")
        except UnicodeDecodeError:
            return raw.decode("latin-1", errors="replace")

    def _ok(
        self,
        text: str,
        fmt: str,
        original_size: int,
        max_bytes: Optional[int],
    ) -> tuple[bool, str]:
        """Truncate *text* and return a success response tuple."""
        text, truncated = self._truncate_text(text, max_bytes)
        return True, json.dumps({
            "content":    text,
            "truncated":  truncated,
            "bytes_read": original_size,
            "format":     fmt,
        })

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------

    def _parse_pdf(
        self, raw: bytes, original_size: int, max_bytes: Optional[int]
    ) -> tuple[bool, str]:
        try:
            doc = fitz.open(stream=raw, filetype="pdf")
            pages_text = [page.get_text() or "" for page in doc]
            doc.close()
            return self._ok("\n\n".join(pages_text), "pdf", original_size, max_bytes)
        except Exception as e:
            return False, json.dumps({"error": f"PDF extraction failed: {e}"})

    # ------------------------------------------------------------------
    # OOXML — docx / xlsx / pptx
    # ------------------------------------------------------------------

    def _parse_ooxml(
        self,
        raw: bytes,
        original_size: int,
        max_bytes: Optional[int],
        fmt: str,
    ) -> tuple[bool, str]:
        # fmt is guaranteed to be one of SUPPORTED_OOXML {"docx", "xlsx", "pptx"}.
        try:
            if fmt == "docx":
                text = DocxParser().parse(BytesIO(raw)).export_to_text()

            elif fmt == "xlsx":
                wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
                rows: list[str] = []
                for ws in wb.worksheets:
                    rows.append(f"=== Sheet: {ws.title} ===")
                    for row in ws.iter_rows(values_only=True):
                        rows.append(
                            "\t".join("" if c is None else str(c) for c in row)
                        )
                text = "\n".join(rows)

            else:  # pptx
                text = PPTXParser().parse_binary(raw).export_to_text()

            return self._ok(text, fmt, original_size, max_bytes)
        except Exception as e:
            return False, json.dumps({"error": f"OOXML extraction failed ({fmt}): {e}"})

    # ------------------------------------------------------------------
    # OLE2 — doc / xls / ppt
    # ------------------------------------------------------------------

    def _parse_ole2(
        self,
        raw: bytes,
        original_size: int,
        max_bytes: Optional[int],
        fmt: str,
    ) -> tuple[bool, str]:
        # fmt is guaranteed to be one of SUPPORTED_OLE2 {"doc", "xls", "ppt"}.
        # All OLE2 formats are handled via converters (LibreOffice) + OOXML parsers.
        try:
            if fmt == "doc":
                text = self._extract_doc_via_parser(raw)
            elif fmt == "xls":
                text = self._extract_xls_via_parser(raw)
            elif fmt == "ppt":
                text = self._extract_ppt_via_parser(raw)
            else:
                return False, json.dumps({"error": f"Unsupported OLE2 format: {fmt}"})
            return self._ok(text, fmt, original_size, max_bytes)
        except Exception as e:
            return False, json.dumps({"error": f"OLE2 extraction failed ({fmt}): {e}"})

    def _extract_doc_via_parser(self, raw: bytes) -> str:
        """Extract text from .doc by converting to .docx (LibreOffice) then parsing."""
        docx_stream = DocParser().convert_doc_to_docx(raw)
        return DocxParser().parse(docx_stream).export_to_text()

    def _extract_xls_via_parser(self, raw: bytes) -> str:
        """Extract text from .xls by converting to .xlsx (LibreOffice) then parsing."""
        xlsx_bytes = XLSParser().convert_xls_to_xlsx(raw)
        wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        rows: list[str] = []
        for ws in wb.worksheets:
            rows.append(f"=== Sheet: {ws.title} ===")
            for row in ws.iter_rows(values_only=True):
                rows.append(
                    "\t".join("" if c is None else str(c) for c in row)
                )
        return "\n".join(rows)

    def _extract_ppt_via_parser(self, raw: bytes) -> str:
        """Extract text from .ppt by converting to .pptx (LibreOffice) then parsing."""
        pptx_bytes = PPTParser().convert_ppt_to_pptx(raw)
        return PPTXParser().parse_binary(pptx_bytes).export_to_text()

    # ------------------------------------------------------------------
    # Plain text — json, html, md, csv, code files, …
    # ------------------------------------------------------------------

    def _parse_plain_text(
        self,
        raw: bytes,
        original_size: int,
        max_bytes: Optional[int],
        ext: str,
    ) -> tuple[bool, str]:
        # Truncate raw bytes before decoding to avoid decoding a huge buffer
        truncated_by_bytes = False
        if max_bytes and len(raw) > max_bytes:
            raw = raw[:max_bytes]
            truncated_by_bytes = True

        text = self._decode_bytes(raw)
        fmt_label = ext  # ext is always a known SUPPORTED_TEXT member at this point

        return True, json.dumps({
            "content":    text,
            "truncated":  truncated_by_bytes,
            "bytes_read": original_size,
            "format":     fmt_label,
        })