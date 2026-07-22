#!/usr/bin/env python3
"""Best-effort static scan for common Excel formula footguns, since this
environment has no way to actually recalculate a workbook (no LibreOffice).
Does NOT execute formulas — only inspects formula text and any cached
values already stored in the file.

Usage: python verify_formulas.py <file.xlsx>
Exit code 0 if clean, 1 if any issue was flagged (details printed to stdout).
"""
from __future__ import annotations

import re
import sys

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

_ERROR_TOKENS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}
_RANGE_RE = re.compile(r"([A-Za-z_][A-Za-z0-9_.]*!)?(\$?[A-Z]{1,3}\$?\d+):(\$?[A-Z]{1,3}\$?\d+)")
_DIVISION_RE = re.compile(r"/\s*(\$?[A-Z]{1,3}\$?\d+)\b")


def verify(path: str) -> list[str]:
    issues: list[str] = []
    wb_formulas = load_workbook(path, data_only=False)
    wb_values = load_workbook(path, data_only=True)

    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name] if sheet_name in wb_values.sheetnames else None
        issues.extend(_check_sheet(sheet_name, ws_f, ws_v))
    return issues


def _check_sheet(sheet_name: str, ws_f, ws_v) -> list[str]:
    issues: list[str] = []
    max_row, max_col = ws_f.max_row, ws_f.max_column

    for row in ws_f.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str) or not value.startswith("="):
                if isinstance(value, str) and value.strip() in _ERROR_TOKENS:
                    issues.append(f"{sheet_name}!{cell.coordinate}: cached error value {value!r}")
                continue

            formula = value

            for match in _RANGE_RE.finditer(formula):
                if match.group(1):  # cross-sheet reference — skip, out of scope for this scan
                    continue
                _min_col, min_row, _max_col, max_row_ref = range_boundaries(f"{match.group(2)}:{match.group(3)}")
                if (max_row_ref or 0) > max_row or (_max_col or 0) > max_col:
                    issues.append(
                        f"{sheet_name}!{cell.coordinate}: formula {formula!r} references range "
                        f"{match.group(2)}:{match.group(3)} which extends beyond the sheet's used "
                        f"range (max_row={max_row}, max_col={max_col}) — check the range was computed "
                        f"from the actual data, not hardcoded"
                    )

            for match in _DIVISION_RE.finditer(formula):
                denom_ref = match.group(1)
                denom_value = _cell_value(ws_v, denom_ref) if ws_v is not None else None
                if denom_value in (0, None, ""):
                    issues.append(
                        f"{sheet_name}!{cell.coordinate}: formula {formula!r} divides by {denom_ref}, "
                        f"whose current value is {denom_value!r} — likely #DIV/0! risk"
                    )

    return issues


def _cell_value(ws, ref: str):
    try:
        return ws[ref.replace("$", "")].value
    except (KeyError, IndexError, ValueError):
        return None


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("usage: python verify_formulas.py <file.xlsx>")
    found = verify(sys.argv[1])
    if not found:
        print("No formula issues flagged.")
        sys.exit(0)
    print(f"{len(found)} potential issue(s) flagged (not executed, static scan only):")
    for issue in found:
        print(f"  - {issue}")
    sys.exit(1)
